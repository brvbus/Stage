# -*- coding: utf-8 -*-
"""
Created on Mon Jul 23 14:18:19 2018

@author: L030909
"""
#from openpyxl import Workbook,load_workbook
#wb2 = load_workbook('testxy.xlsx')
#
#book = Workbook()
#sheet = book.active

#sheet['A1'] = 1
#sheet['B2'] = 2
#sheet['C3'] = 3

#sheet.append('abc')
#
#rows = (
#    (88, 46, 57),
#    (89, 38, 12),
#    (23, 59, 78),
#    (56, 21, 98),
#    (24, 18, 43),
#    (34, 15, 67)
#)
#
#for row in rows:
#    sheet.append(row)
#
#book.save('testxy.xlsx')
#
#print(
#for i in rang:
#    while sheet[]!='NaN'



import pandas
from openpyxl import load_workbook

book = load_workbook('testxy.xlsx')
writer = pandas.ExcelWriter('testxy.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

writer.save()